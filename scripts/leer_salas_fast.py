"""Lee SALAS con openpyxl read_only=True - más rápido."""
import openpyxl, warnings
warnings.filterwarnings('ignore')

EXCEL = r'C:\Users\mario_481\qlik-mcp\Ejecucion Presupuesto Comercial 08 de Mayo de 2026.xlsx'
OUT   = r'C:\Users\mario_481\qlik-mcp\salas_out.txt'

wb = openpyxl.load_workbook(EXCEL, data_only=True, read_only=True)
print(f'Hojas: {wb.sheetnames}')
ws = wb['SALAS']

lines = []
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    if i > 120: break
    vals = list(row[:12])
    if any(v is not None for v in vals):
        lines.append(f'F{i:3d}: {vals}')
        print(f'F{i:3d}: {vals}', flush=True)

wb.close()
with open(OUT, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))
print('DONE', flush=True)
