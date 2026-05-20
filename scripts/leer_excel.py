import openpyxl
from datetime import datetime, date

path = r'C:\Users\mario_481\Downloads\Ejecucion Presupuesto Comercial 30 de abril de 2026.xlsx'
wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
FECHA_30 = date(2026, 4, 30)
DIM2_HOMENAJES = {'2-1','2-2','2-3','2-5','2-6','2-9','2-11','2-13'}

print('=== Otros Ingresos actual — registros del 30/04/2026 ===\n')
ws = wb['Otros Ingresos actual']
headers = None
total_cont = 0
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    if i == 1:
        headers = list(row)
        continue
    if all(c is None for c in row):
        continue
    try:
        fv = row[2]
        if isinstance(fv, datetime): fv = fv.date()
        if fv != FECHA_30: continue
        dim2  = str(row[17]) if row[17] else ''
        nota  = str(row[19]) if row[19] else '0'
        vlr   = float(row[13]) if row[13] else 0
        elem  = str(row[5]) if row[5] else ''
        sede  = str(row[18]) if row[18] else ''
        # Solo sedes homenajes (excluir PREVISIÓN dim2=3-x) y sin nota credito
        if not dim2.startswith('2-') and dim2 != '1-1': continue
        if nota not in ('0','None',''): continue
        print(f'  Dim2: {dim2:<6} Sede: {sede:<15} Elem: {elem:<30} Valor: {vlr:>12,.0f}')
        total_cont += vlr
    except: continue
print(f'\n  TOTAL Contabilidad 30/04 (Homenajes): $ {total_cont:>12,.0f}')

print()
print('=== Facturas directo parque actual — registros del 30/04/2026 ===\n')
ws2 = wb['Facturas directo parque actual']
total_parq = 0
for i, row in enumerate(ws2.iter_rows(values_only=True), 1):
    if i == 1: continue
    if all(c is None for c in row): continue
    try:
        fv = row[2]
        if isinstance(fv, datetime): fv = fv.date()
        if fv != FECHA_30: continue
        nota  = str(row[19]) if len(row) > 19 and row[19] else '0'
        if nota not in ('0','None',''): continue
        dim2  = str(row[17]) if len(row) > 17 and row[17] else ''
        elem  = str(row[5]) if row[5] else ''
        art   = str(row[6]) if row[6] else ''
        vlr   = float(row[13]) if row[13] else 0
        print(f'  Dim2: {dim2:<6} Elem: {elem:<20} Art: {art[:35]:<35} Valor: {vlr:>10,.0f}')
        total_parq += vlr
    except: continue
print(f'\n  TOTAL Facturas Parque 30/04: $ {total_parq:>12,.0f}')

print()
print(f'GRAN TOTAL fuentes 30/04: $ {total_cont+total_parq:>12,.0f}')
wb.close()
