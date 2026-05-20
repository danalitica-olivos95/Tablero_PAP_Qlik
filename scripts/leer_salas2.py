"""Lee hoja SALAS y escribe resultado a archivo."""
import openpyxl, warnings, sys
warnings.filterwarnings('ignore')

EXCEL = r'C:\Users\mario_481\qlik-mcp\Ejecucion Presupuesto Comercial 08 de Mayo de 2026.xlsx'
OUT   = r'C:\Users\mario_481\qlik-mcp\salas_out.txt'

wb = openpyxl.load_workbook(EXCEL, data_only=True)
ws = wb['SALAS']

lines = []
lines.append(f'Hojas: {wb.sheetnames}')
lines.append(f'SALAS: {ws.max_row} filas x {ws.max_column} cols')

for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    if i > 120: break
    vals = list(row[:15])
    if any(v is not None for v in vals):
        lines.append(f'F{i:3d}: {vals}')

with open(OUT, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))

print('OK: salas_out.txt generado')
