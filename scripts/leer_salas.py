"""Lee hoja SALAS del Excel y muestra estructura clave."""
import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
import warnings
warnings.filterwarnings('ignore')

EXCEL = r'C:\Users\mario_481\qlik-mcp\Ejecucion Presupuesto Comercial 08 de Mayo de 2026.xlsx'
wb = openpyxl.load_workbook(EXCEL, data_only=True)
print('Hojas:', wb.sheetnames)
ws = wb['SALAS']

print(f'\n--- SALAS ({ws.max_row} filas x {ws.max_column} cols) ---')
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    if i > 120: break
    vals = list(row[:15])
    if any(v is not None for v in vals):
        print(f'F{i:3d}: {vals}')
