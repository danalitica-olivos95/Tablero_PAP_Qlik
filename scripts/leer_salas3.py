"""Lee hoja SALAS del Excel - output a archivo."""
import openpyxl, warnings
warnings.filterwarnings('ignore')

EXCEL = r'C:\Users\mario_481\qlik-mcp\Ejecucion Presupuesto Comercial 08 de Mayo de 2026.xlsx'
OUT   = r'C:\Users\mario_481\qlik-mcp\salas_out.txt'

wb = openpyxl.load_workbook(EXCEL, data_only=True)
ws = wb['SALAS']

lines = [f'Hojas: {wb.sheetnames}', f'SALAS: {ws.max_row}f x {ws.max_column}c']
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    if i > 120: break
    vals = list(row[:12])
    if any(v is not None for v in vals):
        lines.append(f'F{i:3d}: {vals}')

with open(OUT, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))
print('DONE')
