import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from datetime import datetime, date

path = r'C:\Users\mario_481\Downloads\Ejecucion Presupuesto Comercial 30 de abril de 2026.xlsx'
wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
FECHA_30 = date(2026, 4, 30)

# --- Otros Ingresos actual: ver los 2 registros del 30/04 sin filtro de dim2 ---
print('=== Otros Ingresos actual — registros 30/04 (sin filtro dim2) ===')
ws = wb['Otros Ingresos actual']
headers = None
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    if i == 1:
        headers = list(row)
        # Mostrar headers con índice
        for j, h in enumerate(headers):
            if h: print(f'  Col {j}: {h}')
        print()
        continue
    if all(c is None for c in row): continue
    fv = row[2]
    if isinstance(fv, datetime): fv = fv.date()
    if fv != FECHA_30: continue
    dim2 = str(row[17]) if row[17] else ''
    nota = str(row[19]) if row[19] else ''
    vlr  = float(row[13]) if row[13] else 0
    elem = str(row[5]) if row[5] else ''
    sede = str(row[18]) if row[18] else ''
    print(f'  Dim2: {dim2:<8} Sede: {sede:<15} Elem: {elem:<25} Nota: {nota:<10} Valor: {vlr:>12,.0f}')

print()

# --- Facturas directo parque: inspeccionar cabeceras ---
print('=== Facturas directo parque actual — cabeceras ===')
ws2 = wb['Facturas directo parque actual']
for i, row in enumerate(ws2.iter_rows(values_only=True), 1):
    if i == 1:
        for j, h in enumerate(row):
            if h: print(f'  Col {j}: {h}')
        break

print()
print('=== Facturas directo parque actual — primeras 5 filas (cols fecha-like) ===')
for i, row in enumerate(ws2.iter_rows(values_only=True), 1):
    if i == 1: continue
    if all(c is None for c in row): continue
    if i > 7: break
    # Mostrar columnas 0-5 y algunas más para ver dónde está la fecha
    print(f'  Fila {i}: ' + ' | '.join(str(row[j])[:20] if j < len(row) else '-' for j in [0,1,2,3,4,5,6,13,17,19]))

wb.close()
