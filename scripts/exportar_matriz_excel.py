import asyncio, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.path.insert(0, 'src')
from qlik_mcp.engine_client import SessionManager
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from datetime import date, timedelta
from collections import defaultdict

SEDES_ORDEN = ['CHICO','PALERMO','RESTREPO','SAN DIEGO','TEUSAQUILLO','JPCLO','SOACHA','SOGAMOSO','CHIA','RED OLIVOS']
SEDES_MAP_ID = {'1':'CHICO','2':'PALERMO','3':'RESTREPO','4':'SAN DIEGO',
                '5':'TEUSAQUILLO','6':'JPCLO','7':'SOACHA','8':'SOGAMOSO','10':'CHIA','12':'RED OLIVOS'}

# Estilos
HDR_FILL   = PatternFill("solid", fgColor="1F4E79")
HDR_FONT   = Font(color="FFFFFF", bold=True, size=10)
TOTAL_FILL = PatternFill("solid", fgColor="BDD7EE")
TOTAL_FONT = Font(bold=True, size=10)
GRAY_FILL  = PatternFill("solid", fgColor="F2F2F2")
BORDER_THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'))
NUM_FMT = '#,##0'

def hdr(ws, row, col, value, fill=None, font=None, align='center'):
    c = ws.cell(row=row, column=col, value=value)
    c.fill  = fill or HDR_FILL
    c.font  = font or HDR_FONT
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    c.border = BORDER_THIN
    return c

def val(ws, row, col, value, fmt=NUM_FMT, fill=None, bold=False):
    c = ws.cell(row=row, column=col, value=value)
    c.number_format = fmt
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.border = BORDER_THIN
    if fill: c.fill = fill
    if bold: c.font = Font(bold=True, size=10)
    return c

async def fetch_all_pages(s, sheet_def, width, total_rows):
    """Fetch all pages of a hypercube."""
    handle = await s.create_session_object(sheet_def)
    data = []
    page_size = 500
    offset = 0
    while offset < total_rows:
        result = await s.call(handle, 'GetHyperCubeData',
            ['/qHyperCubeDef', [{'qLeft':0,'qTop':offset,'qWidth':width,'qHeight':page_size}]])
        pages = result.get('qDataPages', [])
        if not pages: break
        rows = pages[0].get('qMatrix', [])
        if not rows: break
        data.extend(rows)
        offset += len(rows)
        if len(rows) < page_size: break
    return data

async def main():
    app_id = '6094b586-82ff-4dd1-9a0e-45c4eb67ef16'
    sessions = SessionManager()
    s = await sessions.get(app_id)

    # --- Obtener vFechaMax ---
    r = await s.call(s.app_handle, 'EvaluateEx', ['=$(vFechaMax)'])
    fecha_max_txt = r.get('qValue', {}).get('qText', '30/04/2026')
    print(f'vFechaMax = {fecha_max_txt}')

    # Construir rango mes actual (01/MM/YYYY .. vFechaMax)
    parts = fecha_max_txt.split('/')
    mes, anio = int(parts[1]), int(parts[2])
    fecha_ini = f'01/{mes:02d}/{anio}'
    fecha_fin = fecha_max_txt

    # Generar lista de fechas del mes para el set analysis
    d0 = date(anio, mes, 1)
    d1 = date(anio, mes, int(parts[0]))
    fechas_lista = []
    d = d0
    while d <= d1:
        fechas_lista.append(d.strftime('%d/%m/%Y'))
        d += timedelta(days=1)
    fechas_set = ','.join(f"'{f}'" for f in fechas_lista)

    # Incluye ADICIONALES JPCLO para capturar la ejecucion real de JPCLO
    ejec_filter = "Ejecucion={'ADICIONALES','ADICIONALES JPCLO'}"
    expr_ejec = f"SUM({{<{ejec_filter}, Fecha={{{fechas_set}}}>}} Ejecutado)"
    expr_ppto = f"SUM({{<Ejecucion={{'ADICIONALES'}}, TipoPresupuesto={{'I'}}, Fecha={{'{fecha_fin}'}}>}} Presupuesto)"

    # ---------------------------------------------------------------
    # HOJA 1: DIARIO — Fecha x Sede
    # ---------------------------------------------------------------
    print('Consultando datos diarios...')
    sheet_diario = {
        'qInfo': {'qType': 'SessionObject'},
        'qHyperCubeDef': {
            'qDimensions': [
                {'qDef': {'qFieldDefs': ['Fecha']}},
                {'qDef': {'qFieldDefs': ['Id_Sede']}},
            ],
            'qMeasures': [{'qDef': {'qDef': expr_ejec, 'qLabel': 'Ejecutado'}}],
            'qSuppressMissing': True,
            'qSuppressZero': True,
            'qInitialDataFetch': [{'qLeft':0,'qTop':0,'qWidth':3,'qHeight':500}]
        }
    }
    handle_d = await s.create_session_object(sheet_diario)
    layout_d = await s.get_layout(handle_d)
    size_d = layout_d.get('qHyperCube',{}).get('qSize',{})
    total_rows_d = size_d.get('qcy', 0)
    print(f'  Filas diario: {total_rows_d}')

    rows_d = await fetch_all_pages(s, sheet_diario, 3, total_rows_d)

    # Organizar: {fecha_str: {sede: valor}}
    diario = defaultdict(dict)
    for row in rows_d:
        f_txt = row[0].get('qText', '')
        id_s  = str(int(row[1]['qNum'])) if row[1].get('qIsNumeric') else row[1].get('qText','')
        sede  = SEDES_MAP_ID.get(id_s, f'Id={id_s}')
        ejec  = row[2].get('qNum', 0) * 1000
        diario[f_txt][sede] = ejec

    # Ordenar fechas
    def parse_fecha(s):
        try:
            p = s.split('/')
            return date(int(p[2]), int(p[1]), int(p[0]))
        except: return date(1900,1,1)
    fechas_ordenadas = sorted(diario.keys(), key=parse_fecha)

    # ---------------------------------------------------------------
    # HOJA 2: MENSUAL — Sede x Categoría de ingreso
    # ---------------------------------------------------------------
    print('Consultando datos mensuales por categoria...')
    sheet_mens = {
        'qInfo': {'qType': 'SessionObject'},
        'qHyperCubeDef': {
            'qDimensions': [
                {'qDef': {'qFieldDefs': ['Id_Sede']}},
            ],
            'qMeasures': [
                {'qDef': {'qDef': expr_ejec, 'qLabel': 'Adicionales'}},
                {'qDef': {'qDef': expr_ppto, 'qLabel': 'Ppto'}},
            ],
            'qSuppressMissing': True,
            'qSuppressZero': True,
            'qInitialDataFetch': [{'qLeft':0,'qTop':0,'qWidth':3,'qHeight':30}]
        }
    }
    handle_m = await s.create_session_object(sheet_mens)
    layout_m = await s.get_layout(handle_m)
    rows_m = layout_m.get('qHyperCube',{}).get('qDataPages',[{}])[0].get('qMatrix',[])

    mensual = {}
    for row in rows_m:
        id_s  = str(int(row[0]['qNum'])) if row[0].get('qIsNumeric') else row[0].get('qText','')
        sede  = SEDES_MAP_ID.get(id_s, f'Id={id_s}')
        ejec  = row[1].get('qNum',0)*1000
        ppto  = row[2].get('qNum',0)*1000
        mensual[sede] = {'ejec': ejec, 'ppto': ppto}

    await sessions.close_all()

    # ---------------------------------------------------------------
    # CREAR EXCEL
    # ---------------------------------------------------------------
    wb = openpyxl.Workbook()

    # ---- HOJA DIARIO ----
    ws_d = wb.active
    ws_d.title = 'ADICIONALES DIARIO'

    # Sedes presentes en los datos (ordenadas)
    sedes_presentes = [s for s in SEDES_ORDEN if any(s in diario[f] for f in fechas_ordenadas)]
    ncols = len(sedes_presentes)

    # Fila 1: título
    ws_d.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols+2)
    c = ws_d.cell(row=1, column=1, value=f'ADICIONALES — Ejecución Diaria  ({fecha_ini} al {fecha_fin})')
    c.font = Font(bold=True, size=12, color="1F4E79")
    c.alignment = Alignment(horizontal='center')

    # Fila 2: cabeceras
    hdr(ws_d, 2, 1, 'FECHA')
    for j, sede in enumerate(sedes_presentes, 2):
        hdr(ws_d, 2, j, sede)
    hdr(ws_d, 2, ncols+2, 'TOTAL DÍA')

    # Filas de datos
    totales_sede = defaultdict(float)
    gran_total = 0
    for i, fecha in enumerate(fechas_ordenadas, 3):
        # Alternar color de fila
        row_fill = GRAY_FILL if i % 2 == 0 else None
        c = ws_d.cell(row=i, column=1, value=fecha)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = BORDER_THIN
        if row_fill: c.fill = row_fill

        fila_total = 0
        for j, sede in enumerate(sedes_presentes, 2):
            v = diario[fecha].get(sede, 0)
            val(ws_d, i, j, v if v else None, fill=row_fill)
            fila_total += v
            totales_sede[sede] += v

        val(ws_d, i, ncols+2, fila_total if fila_total else None, fill=TOTAL_FILL, bold=True)
        gran_total += fila_total

    # Fila TOTAL
    last_row = len(fechas_ordenadas) + 3
    hdr(ws_d, last_row, 1, 'TOTAL MES', fill=PatternFill("solid", fgColor="1F4E79"))
    for j, sede in enumerate(sedes_presentes, 2):
        val(ws_d, last_row, j, totales_sede[sede], fill=TOTAL_FILL, bold=True)
    val(ws_d, last_row, ncols+2, gran_total, fill=PatternFill("solid", fgColor="BDD7EE"), bold=True)

    # Anchos de columna
    ws_d.column_dimensions['A'].width = 14
    for j in range(2, ncols+3):
        ws_d.column_dimensions[get_column_letter(j)].width = 16
    ws_d.freeze_panes = 'B3'
    ws_d.row_dimensions[1].height = 20
    ws_d.row_dimensions[2].height = 30

    # ---- HOJA MENSUAL ----
    ws_m = wb.create_sheet('ADICIONALES MENSUAL')
    ws_m.merge_cells('A1:D1')
    c = ws_m.cell(row=1, column=1, value=f'ADICIONALES — Resumen Mensual  ({fecha_ini} al {fecha_fin})')
    c.font = Font(bold=True, size=12, color="1F4E79")
    c.alignment = Alignment(horizontal='center')

    for col, txt in enumerate(['SEDE', 'PRESUPUESTO MES', 'EJECUTADO', '% CUMPLIMIENTO'], 1):
        hdr(ws_m, 2, col, txt)

    tot_ppto = tot_ejec = 0
    for i, sede in enumerate(SEDES_ORDEN, 3):
        if sede not in mensual: continue
        d = mensual[sede]
        row_fill = GRAY_FILL if i % 2 == 0 else None
        c = ws_m.cell(row=i, column=1, value=sede)
        c.border = BORDER_THIN; c.alignment = Alignment(horizontal='left', vertical='center')
        if row_fill: c.fill = row_fill
        val(ws_m, i, 2, d['ppto'], fill=row_fill)
        val(ws_m, i, 3, d['ejec'], fill=row_fill)
        pct = d['ejec']/d['ppto'] if d['ppto'] else 0
        cv = ws_m.cell(row=i, column=4, value=pct)
        cv.number_format = '0.0%'; cv.border = BORDER_THIN
        cv.alignment = Alignment(horizontal='right')
        if row_fill: cv.fill = row_fill
        tot_ppto += d['ppto']; tot_ejec += d['ejec']

    last = 3 + len([s for s in SEDES_ORDEN if s in mensual])
    hdr(ws_m, last, 1, 'TOTAL', fill=PatternFill("solid", fgColor="1F4E79"))
    val(ws_m, last, 2, tot_ppto, fill=TOTAL_FILL, bold=True)
    val(ws_m, last, 3, tot_ejec, fill=TOTAL_FILL, bold=True)
    pct_t = tot_ejec/tot_ppto if tot_ppto else 0
    cv = ws_m.cell(row=last, column=4, value=pct_t)
    cv.number_format = '0.0%'; cv.fill = TOTAL_FILL; cv.border = BORDER_THIN
    cv.font = Font(bold=True); cv.alignment = Alignment(horizontal='right')

    ws_m.column_dimensions['A'].width = 18
    ws_m.column_dimensions['B'].width = 20
    ws_m.column_dimensions['C'].width = 20
    ws_m.column_dimensions['D'].width = 16
    ws_m.freeze_panes = 'A3'

    # Guardar
    out_path = r'C:\Users\mario_481\Downloads\Matriz_Adicionales_Qlik_30abril2026_v2.xlsx'
    wb.save(out_path)
    print(f'\nArchivo guardado: {out_path}')
    print(f'  - ADICIONALES DIARIO: {len(fechas_ordenadas)} dias x {len(sedes_presentes)} sedes')
    print(f'  - ADICIONALES MENSUAL: {len([s for s in SEDES_ORDEN if s in mensual])} sedes')
    print(f'  - Gran total mes: $ {gran_total:,.0f}')

asyncio.run(main())
